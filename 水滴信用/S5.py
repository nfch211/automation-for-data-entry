import pickle
import os
import threading
from tkinter import Tk, filedialog
from docx import Document
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import random
import time
from selenium.common.exceptions import NoSuchElementException






# Function to wait for an element to disappear
def wait_for_element_disappearance(driver, xpath, timeout=180):
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            driver.find_element(By.XPATH, xpath)
            print("Waiting for element to disappear...")
            time.sleep(3)
        except NoSuchElementException:
            return True
    return False


# Function to load cookies
def load_cookies(driver, url, cookie_path):
    try:
        driver.get(url)
        cookies = pickle.load(open(cookie_path, "rb"))
        for cookie in cookies:
            driver.add_cookie(cookie)
        print("Cookies loaded successfully.")
        return True
    except FileNotFoundError:
        print("Cookie file not found. You will need to log in.")
        return False

# Function to save cookies
def save_cookies(driver, cookie_path):
    pickle.dump(driver.get_cookies(), open(cookie_path, "wb"))
    print("Cookies saved successfully.")

# Function to check for script exit
def exit_script_check():
    global exit_requested
    input("Press Enter at any time to stop the script and save the workbook.")
    exit_requested = True

# Function for manual user login
def user_login(driver, user_agent, chrome_binary_location, chrome_driver_path):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.binary_location = chrome_binary_location
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_argument("--log-level=3")

    driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)
    driver.get('https://shuidi.cn/')
    print("Please perform the login manually in the web browser that just opened.")
    time.sleep(20)
    print("Once you are logged in successfully, come back to the terminal and press Enter to continue...")
    input()

# Setup
Tk().withdraw()
excel_filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
workbook = openpyxl.load_workbook(excel_filename)
sheet = workbook.active

chrome_binary_location = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
chrome_driver_path = r'C:\Users\hofong\Desktop\python\水滴信用\chromedriver.exe'
cookies_path = r"C:\Users\hofong\Desktop\python\水滴信用\cookies\cookies.pkl"

user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36",
    "Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2486.0 Safari/537.36 Edge/13.10586",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.84 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1"
]

user_agent = random.choice(user_agents)
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = chrome_binary_location
chrome_options.add_argument(f'user-agent={user_agent}')
chrome_options.add_argument("--log-level=3")

# Start Selenium WebDriver
driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)

# Load or Save Cookies
if not load_cookies(driver, 'https://shuidi.cn/', cookies_path):
    user_login(driver, user_agent, chrome_binary_location, chrome_driver_path)
    save_cookies(driver, cookies_path)
    load_cookies(driver, 'https://shuidi.cn/', cookies_path)

# Data Extraction Loop
exit_requested = False
threading.Thread(target=exit_script_check, daemon=True).start()

for row_num in range(2, sheet.max_row + 1):
    if exit_requested:
        break

    br_number = str(sheet.cell(row=row_num, column=1).value).strip()

    if not br_number.startswith(('9', '4', '3', '1')) or '-' in br_number or len(br_number) < 15:
        continue

    if sheet.cell(row=row_num, column=2).value and sheet.cell(row=row_num, column=3).value:
        continue

    data_url = f'https://shuidi.cn/pc-search?key={br_number}'
    driver.get(data_url)

    # Wait for the specific div element to disappear
    xpath_to_check = "//div[@style='margin-left: 6px; float: left;']"
    if not wait_for_element_disappearance(driver, xpath_to_check):
        print("Element did not disappear in the given time.")
        continue

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, 'html.parser')

    company_name = soup.find('a', class_='name_row')
    company_name_str = company_name.text.strip() if company_name else "N/A"
    sheet.cell(row=row_num, column=2).value = company_name_str

    company_address = soup.find('span', class_='text_address text-active')
    company_address_str = company_address.text.strip() if company_address else "N/A"
    sheet.cell(row=row_num, column=3).value = company_address_str

    print(f"BR Number: {br_number}, Company Name: {company_name_str}, Address: {company_address_str}")

    workbook.save(excel_filename)

    docx_path = sheet.cell(row=row_num, column=8).value

    if docx_path and docx_path.endswith('.docx') and os.path.exists(docx_path):
        try:
            doc = Document(docx_path)
            docx_text = '\n'.join([para.text for para in doc.paragraphs])

            blue_color = "0000FF"
            red_color = "FF0000"

            if company_name_str in docx_text:
                sheet.cell(row=row_num, column=2).font = openpyxl.styles.Font(color=blue_color)
            else:
                sheet.cell(row=row_num, column=2).font = openpyxl.styles.Font(color=red_color)

            if company_address_str in docx_text:
                sheet.cell(row=row_num, column=3).font = openpyxl.styles.Font(color=blue_color)
            else:
                sheet.cell(row=row_num, column=3).font = openpyxl.styles.Font(color=red_color)
        except Exception as e:
            print(f"Error reading DOCX file for BR number {br_number}: {e}")
    else:
        print(f"Invalid DOCX path for BR number {br_number}: {docx_path}")

workbook.save(excel_filename)
print("Workbook saved. Exiting script.")
driver.quit()
